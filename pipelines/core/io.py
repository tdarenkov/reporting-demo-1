"""I/O helpers for source-file ingest.

Replaces the per-notebook duplicated `read_*` helpers from the Fabric
notebooks. Each function returns a list of dicts (one per row) so callers
can iterate without touching pandas/openpyxl directly.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def read_titled_csv(path: Path, skip_rows: int = 2) -> pd.DataFrame:
    """Read a CSV that begins with `skip_rows` non-data title rows. All-string columns."""
    return pd.read_csv(path, skiprows=skip_rows, dtype=str)


def read_xlsx_first_sheet(path: Path) -> list[dict[str, Any]]:
    """Read the first worksheet into a list of {header: cell} dicts.

    Skips fully blank rows. Cells return as native Python types (openpyxl
    gives strings for text cells, numbers for numeric cells, etc.) so each
    per-table transform can decide how to coerce.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        out.append({headers[i]: cell for i, cell in enumerate(row) if i < len(headers)})
    return out


def read_named_table(path: Path, table_name: str) -> list[dict[str, Any]]:
    """Read an Excel "Table" object by name from an xlsx workbook.

    The source workbooks expose curated subsets of data as named Excel
    tables (the Insert → Table feature). This walks every worksheet
    looking for the matching table name and returns it as a list of
    {header: cell} dicts.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            ref = ws.tables[table_name].ref
            rows = list(ws[ref])
            if not rows:
                return []
            headers = [c.value for c in rows[0]]
            return [
                {headers[i]: c.value for i, c in enumerate(row) if i < len(headers)}
                for row in rows[1:]
            ]
    raise ValueError(f"Named table {table_name!r} not found in {path}")


def read_xlsx_all_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read every worksheet of an xlsx into {sheet_name: [{header: cell}, ...]}.

    Used by single-workbook multi-sheet sources.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[ws.title] = []
            continue
        headers = list(rows[0])
        out[ws.title] = [
            {headers[i]: c for i, c in enumerate(row) if i < len(headers)}
            for row in rows[1:]
            if not all(cell is None for cell in row)
        ]
    return out


def read_semicolon_csv(path: Path, **kwargs: Any) -> pd.DataFrame:
    """Read a semicolon-delimited CSV (utf-8-sig BOM, NULL string sentinels)."""
    return pd.read_csv(
        path, sep=";", encoding="utf-8-sig", na_values=["NULL"], dtype=str, **kwargs
    )


def read_json(path: Path) -> Any:
    """Read a UTF-8 JSON file into native Python."""
    return json.loads(path.read_text(encoding="utf-8"))


def latest_match(directory: Path, pattern: str, suffix: str = "") -> Path:
    """Return the most-recent file under directory whose name contains pattern.

    Sort is lexicographic. That equals chronological order only when the
    embedded date is zero-padded ISO (`YYYY-MM-DD`) — which is what the
    current callers use. A `MMDDYY` stamp does NOT sort chronologically
    (e.g. `010125` < `020124`), so a caller using `MMDDYY` filenames must
    not rely on this to pick the latest file.
    """
    matches = sorted(
        p for p in directory.iterdir() if pattern in p.name and p.name.endswith(suffix)
    )
    if not matches:
        raise FileNotFoundError(f"No file matching {pattern!r}{suffix} under {directory}")
    return matches[-1]
